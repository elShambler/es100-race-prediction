import base64
import json
import logging
from io import BytesIO
from pathlib import Path

import matplotlib
import polars as pl

matplotlib.use("Agg")  # kedro runs headless; never open a GUI window
import matplotlib.pyplot as plt

from eastern_states_pace_predict import mpl_theme

logger = logging.getLogger(__name__)

RACE_START_HR = 5  # 05:00 start, all years
MIN_GROUP_N = 5  # suppress aggregate cells with fewer runners than this
MINUTES_PER_HOUR = 60
# Stoppage medians are only shown when observed check-in/out pairs cover at
# least this share of a station's visits (guards against the biased subset
# left by sparse 2025 check-in recording).
STOPPAGE_COVERAGE_MIN = 0.3


def _fmt_tod(elapsed_hrs: float) -> str:
    """Clock time for an elapsed-hours value; '+1' marks the second day."""
    total = RACE_START_HR + elapsed_hrs
    day, tod = divmod(total, 24)
    h = int(tod)
    m = int(round((tod - h) * MINUTES_PER_HOUR))
    if m == MINUTES_PER_HOUR:
        h, m = h + 1, 0
    suffix = " +1" if day >= 1 else ""
    return f"{h:02d}:{m:02d}{suffix}"


def _quantiles(s: pl.Series, qs: tuple[float, ...]) -> list[float | None]:
    s = s.drop_nulls()
    if s.len() < MIN_GROUP_N:
        return [None] * len(qs)
    return [round(s.quantile(q, interpolation="linear"), 3) for q in qs]


def _speed_ratio_expr(pace_ratio_col: str) -> pl.Expr:
    """Leg speed relative to the runner's final overall pace.

    Inverts a pace ratio (leg pace ÷ final overall pace) into a speed ratio
    (final overall pace ÷ leg pace), so >1.0 means that leg ran *faster* than
    the runner's whole-race average and <1.0 means slower. Inverting per-row
    before aggregating is required — median(1/x) ≠ 1/median(x). Non-positive
    pace ratios yield null.
    """
    return (
        pl.when(pl.col(pace_ratio_col) > 0)
        .then(1.0 / pl.col(pace_ratio_col))
        .otherwise(None)
    )


def _leg_mph_expr(pace_col: str) -> pl.Expr:
    """Absolute moving speed (mph) from a leg pace in minutes per mile."""
    return pl.when(pl.col(pace_col) > 0).then(60.0 / pl.col(pace_col)).otherwise(None)


def _year_payload(df: pl.DataFrame) -> dict:
    """All chart aggregates for one year's interval features."""
    runners = df.unique(subset=["bib"])
    n_start = runners.height
    n_finish = runners.filter(pl.col("is_finisher")).height
    finish_hrs = (
        runners.filter(pl.col("is_finisher"))["overall_pace_min_per_mi"] * 103.1 / 60
    )

    stations = (
        df.unique(subset=["as_index"])
        .sort("as_index")
        .select(["as_index", "as_name", "as_num", "as_dist_from_start"])
    )

    # Arrival windows: p10/p25/p50/p75/p90 of check-in elapsed hours per station
    # (imputed 2025 check-ins included — they carry the arrival estimate).
    arrivals = []
    # Stoppage: observed (non-imputed) medians, DNF cohort vs finishers.
    stoppage = []
    # Leg difficulty: median pace ratio per leg, clean legs only.
    legs = []
    # Attrition: furthest station reached by DNF runners.
    dnf_by_max = (
        runners.filter(~pl.col("is_finisher"))
        .group_by("MaxAS")
        .len()
        .rename({"MaxAS": "as_index", "len": "n"})
    )
    attrition = []
    # Flow heatmap: arrivals per station per hour of race time.
    heat_rows = []

    for st in stations.iter_rows(named=True):
        sdf = df.filter(pl.col("as_index") == st["as_index"])
        label = f"{st['as_name']}"
        dist = st["as_dist_from_start"]

        q = _quantiles(sdf["as_check_in__elapsed__min"], (0.1, 0.25, 0.5, 0.75, 0.9))
        arrivals.append(
            {
                "as": st["as_index"],
                "name": label,
                "dist": dist,
                "q10": q[0],
                "q25": q[1],
                "q50": q[2],
                "q75": q[3],
                "q90": q[4],
                "tod50": _fmt_tod(q[2]) if q[2] is not None else None,
                "tod10": _fmt_tod(q[0]) if q[0] is not None else None,
                "tod90": _fmt_tod(q[4]) if q[4] is not None else None,
                "n": sdf["as_check_in__elapsed__min"].drop_nulls().len(),
            }
        )

        # Observed pairs only — and only when they cover enough of the
        # station's visits. Years with sparse check-in recording (2025) leave
        # a biased subset at some stations: the few runners whose arrival got
        # recorded are often exactly the ones who stopped long.
        observed = sdf.filter(~pl.col("stoppage_imputed"))
        coverage = observed.height / sdf.height if sdf.height else 0.0
        fin = observed.filter(pl.col("is_finisher"))["as_stoppage_time_min"]
        dnf = observed.filter(~pl.col("is_finisher"))["as_stoppage_time_min"]
        reliable = coverage >= STOPPAGE_COVERAGE_MIN
        stoppage.append(
            {
                "as": st["as_index"],
                "name": label,
                "dist": dist,
                "finisher": _quantiles(fin, (0.5,))[0] if reliable else None,
                "dnf": _quantiles(dnf, (0.5,))[0] if reliable else None,
                "n_finisher": fin.drop_nulls().len(),
                "n_dnf": dnf.drop_nulls().len(),
            }
        )

        # Absolute leg speed (mph): median moving speed on the leg into this
        # station. The per-year card shows this against the year's median leg
        # speed line (the planner card keeps the mph *ratio* instead).
        clean = sdf.filter(~pl.col("spans_missing_as")).with_columns(
            _leg_mph_expr("as_interval_pace").alias("_leg_mph")
        )
        mph = _quantiles(clean["_leg_mph"], (0.5,))[0]
        pace = _quantiles(clean["as_interval_pace"], (0.5,))[0]
        legs.append(
            {
                "as": st["as_index"],
                "name": label,
                "dist": dist,
                "leg_mi": round(clean["interval_dist_mi"].drop_nulls().median() or 0, 1)
                if clean.height
                else None,
                "mph": mph,
                "pace": pace,
                "n": clean["_leg_mph"].drop_nulls().len(),
            }
        )

        n_dnf_here = dnf_by_max.filter(pl.col("as_index") == st["as_index"])
        attrition.append(
            {
                "as": st["as_index"],
                "name": label,
                "dist": dist,
                "n": int(n_dnf_here["n"][0]) if n_dnf_here.height else 0,
            }
        )

        counts = (
            sdf.with_columns(
                pl.col("as_check_in__elapsed__min").floor().cast(pl.Int32).alias("_hr")
            )
            .group_by("_hr")
            .len()
        )
        by_hr = dict(zip(counts["_hr"].to_list(), counts["len"].to_list()))
        heat_rows.append([int(by_hr.get(h, 0)) for h in range(36)])

    med_stop = df.filter(~pl.col("stoppage_imputed"))["as_stoppage_time_min"]

    # Year's median leg speed (mph) — the reference line on the leg-speed card,
    # over every clean moving leg (not the per-station medians).
    year_legs = df.filter(~pl.col("spans_missing_as")).with_columns(
        _leg_mph_expr("as_interval_pace").alias("_leg_mph")
    )
    median_leg_speed = _quantiles(year_legs["_leg_mph"], (0.5,))[0]

    return {
        "kpis": {
            "starters": n_start,
            "finishers": n_finish,
            "finish_rate": round(n_finish / n_start, 3) if n_start else None,
            "median_finish_hrs": round(finish_hrs.median(), 2)
            if finish_hrs.len()
            else None,
            "median_stoppage_min": round(med_stop.drop_nulls().median(), 1)
            if med_stop.drop_nulls().len()
            else None,
        },
        "stations": stations.rows(named=True),
        "arrivals": arrivals,
        "stoppage": stoppage,
        "legs": legs,
        "median_leg_speed": median_leg_speed,
        "attrition": attrition,
        "heat": {"hours": list(range(36)), "rows": heat_rows},
    }


HALF_HOURS_IN_RACE = 72  # 36 h of race time in half-hour arrival bins


def _planner_payload(
    ratio: pl.DataFrame,
    splits: pl.DataFrame,
    xwalk: pl.DataFrame,
    stations: pl.DataFrame,
    params: dict,
) -> dict:
    """Year-independent planner aggregates for the scatter + arrival cards.

    All years pooled, finishers only for the pace-ratio scatter/trend; the
    arrival histogram uses every recorded arrival (DNFs included) so the
    distribution reflects who actually passed through.
    """
    fhr_min = params["finish_hr_min"]
    fhr_max = params["finish_hr_max"]

    # Speed ratio (final pace ÷ leg pace): >1.0 = leg faster than the runner's
    # overall average. Inverted per-row here so all downstream means/samples are
    # computed on the speed metric (median/mean of 1/x ≠ 1/median of x).
    ratio = ratio.with_columns(_speed_ratio_expr("interval_ratio").alias("_speed"))

    # Selectable stations = the 2026 aid stations, Start excluded (nothing runs
    # into it). [id, name, scaled mile].
    selectable = stations.filter(pl.col("station_id") > 0).sort("station_id")
    station_rows = [
        [int(r["station_id"]), r["name"], round(r["scaled_mi"], 1)]
        for r in selectable.iter_rows(named=True)
    ]

    # Scatter points: one per finisher × station, mapped to a 2026 station.
    # Stratified-sample down to max_scatter_points, keeping cohorts balanced.
    pts_df = ratio.filter(
        pl.col("station_2026").is_not_null() & pl.col("_speed").is_not_null()
    )
    cap = params["max_scatter_points"]
    if pts_df.height > cap:
        frac = cap / pts_df.height
        pts_df = (
            pts_df.with_columns(pl.col("finish_hr_block").alias("_blk"))
            .filter(
                pl.int_range(pl.len()).shuffle(seed=17).over("_blk")
                < (pl.len().over("_blk") * frac).ceil()
            )
            .drop("_blk")
        )
    points = [
        [
            round(r["as_dist_from_start"], 2),
            round(r["_speed"], 3),
            int(r["station_2026"]),
            int(r["finish_hr_block"]),
        ]
        for r in pts_df.iter_rows(named=True)
    ]

    # Per-station average speed ratio (full data, not the sample):
    # {sid: [mean, n]}.
    avg = {}
    avg_df = (
        ratio.filter(pl.col("station_2026").is_not_null())
        .group_by("station_2026")
        .agg(pl.col("_speed").mean().alias("m"), pl.len().alias("n"))
    )
    for r in avg_df.iter_rows(named=True):
        avg[str(int(r["station_2026"]))] = [round(r["m"], 3), int(r["n"])]

    # Cohort trend: mean speed ratio per (finish-hour block, station); cells with
    # fewer than MIN_GROUP_N runners are dropped so a lone runner can't define a
    # "trend". {fhr: [[sid, mean, n], ...]}.
    trend = {}
    trend_df = (
        ratio.filter(
            pl.col("station_2026").is_not_null()
            & (pl.col("finish_hr_block") >= fhr_min)
            & (pl.col("finish_hr_block") <= fhr_max)
        )
        .group_by("finish_hr_block", "station_2026", "station_mi_2026")
        .agg(pl.col("_speed").mean().alias("m"), pl.len().alias("n"))
        .filter(pl.col("n") >= MIN_GROUP_N)
        .sort("station_mi_2026")
    )
    for r in trend_df.iter_rows(named=True):
        trend.setdefault(str(int(r["finish_hr_block"])), []).append(
            [int(r["station_2026"]), round(r["m"], 3), int(r["n"])]
        )

    # Arrival distributions: every recorded arrival, mapped to a 2026 station via
    # the crosswalk. Elapsed hours live in as_check_in__elapsed__min (decimal
    # hours despite the name); bin into half-hours over 0–36 h.
    arr = (
        splits.select(["year", "as_index", "as_check_in__elapsed__min", "FinishRank"])
        .join(
            xwalk.select(["year", "as_index", "station_2026"]),
            on=["year", "as_index"],
            how="inner",
        )
        .filter(
            pl.col("station_2026").is_not_null()
            & pl.col("as_check_in__elapsed__min").is_not_null()
        )
        .with_columns(pl.col("as_check_in__elapsed__min").alias("hrs"))
    )
    bins: dict[str, list[int]] = {}
    bin_df = (
        arr.with_columns(
            (pl.col("hrs") * 2)
            .floor()
            .cast(pl.Int32)
            .clip(0, HALF_HOURS_IN_RACE - 1)
            .alias("b")
        )
        .group_by("station_2026", "b")
        .len()
    )
    for r in bin_df.iter_rows(named=True):
        sid = str(int(r["station_2026"]))
        bins.setdefault(sid, [0] * HALF_HOURS_IN_RACE)[int(r["b"])] = int(r["len"])

    # Cohort arrival window: p25/p50/p75 arrival hour at each station for the
    # finishers in each finish-hour block (from the finishers-only ratio frame,
    # which already carries elapsed_hrs + finish_hr_block per station).
    cohort: dict[str, dict[str, list]] = {}
    for r in (
        ratio.filter(
            pl.col("station_2026").is_not_null()
            & (pl.col("finish_hr_block") >= fhr_min)
            & (pl.col("finish_hr_block") <= fhr_max)
        )
        .group_by("station_2026", "finish_hr_block")
        .agg(
            pl.col("elapsed_hrs").quantile(0.25, "linear").alias("p25"),
            pl.col("elapsed_hrs").quantile(0.5, "linear").alias("p50"),
            pl.col("elapsed_hrs").quantile(0.75, "linear").alias("p75"),
            pl.len().alias("n"),
        )
        .iter_rows(named=True)
    ):
        if r["n"] < MIN_GROUP_N:
            continue
        sid = str(int(r["station_2026"]))
        cohort.setdefault(sid, {})[str(int(r["finish_hr_block"]))] = [
            round(r["p25"], 2),
            round(r["p50"], 2),
            round(r["p75"], 2),
            int(r["n"]),
        ]

    return {
        "stations": station_rows,
        "fhr_min": fhr_min,
        "fhr_max": fhr_max,
        "points": points,
        "avg": avg,
        "trend": trend,
        "arrivals": {"bins": bins, "cohort": cohort},
    }


def _course_payload(route: pl.DataFrame, stations: pl.DataFrame, params: dict) -> dict:
    """Downsampled 2026 route + station markers for the Leaflet map card.

    Route is thinned to max_route_points while always keeping the station
    vertices, so each station's routeIdx indexes cleanly into the kept array.
    """
    dec = params["coord_decimals"]
    cap = params["max_route_points"]
    n = route.height
    keep_seqs = set(stations["route_seq"].to_list())
    step = max(1, (n + cap - 1) // cap)
    kept = [
        r
        for i, r in enumerate(route.sort("seq").iter_rows(named=True))
        if i % step == 0 or r["seq"] in keep_seqs
    ]
    seq_to_idx = {r["seq"]: i for i, r in enumerate(kept)}
    route_pts = [[round(r["lat"], dec), round(r["lon"], dec)] for r in kept]

    station_rows = []
    for r in stations.sort("station_id").iter_rows(named=True):
        station_rows.append(
            [
                int(r["station_id"]),
                r["name"],
                round(r["lat"], dec),
                round(r["lon"], dec),
                round(r["scaled_mi"], 1),
                seq_to_idx[r["route_seq"]],
            ]
        )
    total_mi = round(
        stations.filter(pl.col("station_id") == pl.col("station_id").max())[
            "scaled_mi"
        ][0],
        1,
    )
    return {"route": route_pts, "stations": station_rows, "total_mi": total_mi}


def build_as_dashboard(
    features: pl.DataFrame,
    ratio: pl.DataFrame,
    splits: pl.DataFrame,
    route: pl.DataFrame,
    stations: pl.DataFrame,
    xwalk: pl.DataFrame,
    params: dict,
) -> str:
    """Render the aid-station dashboard as a self-contained HTML page.

    Aggregates es_interval_features per year and injects the JSON into the
    HTML/CSS/JS template that lives next to this module. Everything except the
    Leaflet map card is self-contained; the map pulls OpenStreetMap tiles over
    the network at view time.

    Inputs: es_interval_features, es_interval_ratio, es_splits_all,
        es_course_route, es_course_stations, es_station_xwalk, params:reporting
    Outputs: es_as_dashboard (text HTML, data/08_reporting)
    """
    payload = {
        "years": sorted(features["year"].unique().to_list(), reverse=True),
        "generated_note": "Eastern States 100 — split data 2016–2025",
        "by_year": {
            str(y): _year_payload(features.filter(pl.col("year") == y))
            for y in sorted(features["year"].unique().to_list())
        },
        "planner": _planner_payload(ratio, splits, xwalk, stations, params),
        "course": _course_payload(route, stations, params),
    }

    template = (Path(__file__).parent / "template.html").read_text(encoding="utf-8")
    marker = "/*__DATA__*/null"
    if marker not in template:
        raise ValueError("dashboard template is missing the /*__DATA__*/ marker")
    # </script> inside a JSON string would end the script block early.
    blob = json.dumps(payload, separators=(",", ":")).replace("</", "<\\/")
    html = template.replace(marker, blob)
    logger.info(
        "Dashboard built: %d years, %d bytes", len(payload["by_year"]), len(html)
    )
    return html


def _interp_by_mile(miles: list, arr: list) -> None:
    """Fill None entries of arr in place by linear interpolation on mile.

    Interior gaps interpolate between the nearest filled neighbors; leading /
    trailing gaps scale the nearest filled value by the mile ratio, so a station
    no cohort ever recorded still gets a distance-sensible arrival estimate.
    """
    n = len(arr)
    known = [i for i in range(n) if arr[i] is not None]
    if not known:
        for i in range(n):
            arr[i] = miles[i] or 0.0
        return
    first, lastk = known[0], known[-1]
    for i in range(first):
        arr[i] = arr[first] * (miles[i] / miles[first]) if miles[first] else arr[first]
    for i in range(lastk + 1, n):
        arr[i] = arr[lastk] * (miles[i] / miles[lastk]) if miles[lastk] else arr[lastk]
    for a, b in zip(known, known[1:]):
        if b - a > 1:
            span = miles[b] - miles[a]
            for i in range(a + 1, b):
                t = (miles[i] - miles[a]) / span if span else 0.0
                arr[i] = arr[a] + t * (arr[b] - arr[a])


def _monotone(seq: list) -> list:
    """Running max — arrivals never move backward as distance increases."""
    out, m = [], float("-inf")
    for x in seq:
        m = max(m, x)
        out.append(m)
    return out


def _planner_eta_table(planner: dict) -> dict:
    """Per finish-hour-cohort arrival-fraction table for the pacing planner.

    For each finish-hour block b in [fhr_min, fhr_max] and each 2026 station
    (sorted by mile), take that cohort's **median arrival elapsed hours** — the
    nearest block when b itself never recorded the station (nearest-neighbor, not
    a global average) — plus p25/p75, and divide by the block's finish-station
    median so every value is a *fraction of the finish*. The client and the
    workbook multiply these by the user's goal, so the predicted finish equals
    the goal exactly and each ETA carries the cohort's real arrival shape
    (aid-station stoppage included, unlike a moving-pace model). Stations no
    cohort ever recorded are interpolated by mile. Returns
    {str(block): {"p25": [...], "p50": [...], "p75": [...]}} aligned to the
    mile-sorted stations.
    """
    stations = sorted(planner["stations"], key=lambda r: r[2])
    sids = [int(s[0]) for s in stations]
    miles = [s[2] for s in stations]
    cohort = planner["arrivals"]["cohort"]  # {sid: {block: [p25, p50, p75, n]}}
    fmin, fmax = int(planner["fhr_min"]), int(planner["fhr_max"])

    avail = {
        sid: sorted((int(b), v[:3]) for b, v in cohort.get(str(sid), {}).items())
        for sid in sids
    }

    def nearest(sid: int, b: int):
        lst = avail[sid]
        return min(lst, key=lambda kv: (abs(kv[0] - b), kv[0]))[1] if lst else None

    eta = {}
    for b in range(fmin, fmax + 1):
        cols = {q: [None] * len(sids) for q in ("p25", "p50", "p75")}
        for i, sid in enumerate(sids):
            v = nearest(sid, b)
            if v is not None:
                cols["p25"][i], cols["p50"][i], cols["p75"][i] = v
        for q in cols:
            _interp_by_mile(miles, cols[q])
        mfin = cols["p50"][-1] or 1.0
        # Cap at the finish (1.0) so nearest-neighbour mixing across cohorts can't
        # place a median arrival "after" the finish; the finish stays exactly 1.0
        # and the predicted finish lands on the goal.
        f50 = _monotone([min(x / mfin, 1.0) for x in cols["p50"]])
        f25 = [min(x / mfin, m) for x, m in zip(cols["p25"], f50)]
        f75 = [max(x / mfin, m) for x, m in zip(cols["p75"], f50)]
        eta[str(b)] = {
            "p25": [round(x, 4) for x in f25],
            "p50": [round(x, 4) for x in f50],
            "p75": [round(x, 4) for x in f75],
        }
    return eta


def build_planner_workbook(  # noqa: PLR0915
    stations: list, eta: dict, fmin: int, fmax: int
) -> bytes:
    """A self-contained .xlsx pacing planner with the same live logic as the page.

    Enter a goal finish (hours) in Plan!B1 and every predicted arrival recomputes;
    type actual arrivals (elapsed hours) in the Actual column and the remaining
    stations + projected finish re-project from the furthest actual — exactly like
    renderPacing / recomputePacing in planner_template.html. Predicted arrivals are
    the finish-cohort's arrival fractions (baked into a hidden Data sheet, keyed by
    station row × finish-hour block) times the goal, so no macros are needed and it
    works offline. Returns the workbook as bytes.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter as col

    stations = sorted(stations, key=lambda r: r[2])
    n = len(stations)
    total = stations[-1][2]
    blocks = list(range(fmin, fmax + 1))
    nb = len(blocks)

    wb = Workbook()

    # ---- hidden Data sheet: arrival-fraction matrices (p50/p25/p75) ----
    data = wb.active
    data.title = "Data"
    C50, C25, C75 = 4, 4 + nb + 1, 4 + 2 * (nb + 1)  # A name | B mile | (C gap) | …
    for c0 in (C50, C25, C75):  # block-number header per region
        for j, b in enumerate(blocks):
            data.cell(row=1, column=c0 + j, value=b)
    for i, (_sid, name, mi) in enumerate(stations):
        r = i + 2
        data.cell(row=r, column=1, value=name)
        data.cell(row=r, column=2, value=round(mi, 2))
        for j, b in enumerate(blocks):
            eb = eta[str(b)]
            data.cell(row=r, column=C50 + j, value=eb["p50"][i])
            data.cell(row=r, column=C25 + j, value=eb["p25"][i])
            data.cell(row=r, column=C75 + j, value=eb["p75"][i])
    data.sheet_state = "hidden"

    a50, z50 = col(C50), col(C50 + nb - 1)
    a25, z25 = col(C25), col(C25 + nb - 1)
    a75, z75 = col(C75), col(C75 + nb - 1)

    # ---- Plan sheet: user-facing, all formulas ----
    plan = wb.create_sheet("Plan")
    first, last = 5, n + 4  # data rows
    goal = "$B$1"
    # round the goal to a finish-hour block, clamped into the cohort range.
    blk = f"MEDIAN({fmin},ROUND({goal},0),{fmax})"

    def clk(expr: str) -> str:  # elapsed-hours expr -> "hh:mm" (+1 next day)
        return (
            f'TEXT(MOD((5+({expr}))/24,1),"hh:mm")&'
            f'IF(INT((5+({expr}))/24)>=1," +1","")'
        )

    plan["A1"] = "Goal finish (hours)"
    plan["A1"].font = Font(bold=True)
    plan["B1"] = 28
    plan["B1"].number_format = "0.0"
    plan["B1"].fill = PatternFill("solid", fgColor="FFF3D6")
    plan["A2"] = "Projected finish"
    plan["A2"].font = Font(bold=True)
    plan["B2"] = f"=IF($L$1>0,$L$5+({goal}-$L$4)*$L$6,{goal})"
    plan["B2"].number_format = "0.00"
    plan["C2"] = "=" + clk("$B$2")
    plan["D2"] = "min vs goal"
    plan["E2"] = f"=($B$2-{goal})*60"
    plan["E2"].number_format = "+0;-0"

    headers = [
        "Aid station",
        "Mile",
        "Predicted (hrs)",
        "Predicted",
        "Typical range",
        "Actual (elapsed hrs)",
        "Revised (hrs)",
        "Revised",
    ]
    for c, htxt in enumerate(headers, start=1):
        cell = plan.cell(row=4, column=c, value=htxt)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="right" if c > 1 else "left")

    for i, (_sid, name, mi) in enumerate(stations):
        r, dr = first + i, i + 2
        row50, hdr50 = f"Data!${a50}{dr}:${z50}{dr}", f"Data!${a50}$1:${z50}$1"
        row25, hdr25 = f"Data!${a25}{dr}:${z25}{dr}", f"Data!${a25}$1:${z25}$1"
        row75, hdr75 = f"Data!${a75}{dr}:${z75}{dr}", f"Data!${a75}$1:${z75}$1"
        v50 = f"INDEX({row50},MATCH({blk},{hdr50},0))*{goal}"
        v25 = f"INDEX({row25},MATCH({blk},{hdr25},0))*{goal}"
        v75 = f"INDEX({row75},MATCH({blk},{hdr75},0))*{goal}"

        plan.cell(row=r, column=1, value=name)
        plan.cell(row=r, column=2, value=round(mi, 2))
        # Predicted arrival hrs = nearest finish-cohort arrival fraction * goal.
        plan.cell(
            row=r, column=3, value=f"=IFERROR({v50},{goal}*B{r}/{total})"
        ).number_format = "0.00"
        plan.cell(row=r, column=4, value="=" + clk(f"C{r}"))
        plan.cell(row=r, column=5, value=f'=IFERROR({clk(v25)}&" – "&{clk(v75)},"—")')
        act = plan.cell(row=r, column=6)  # input
        act.fill = PatternFill("solid", fgColor="FFF3D6")
        act.number_format = "0.00"
        # Revised hrs: actual if entered, else re-project from the furthest actual.
        plan.cell(
            row=r,
            column=7,
            value=f"=IF(ISNUMBER(F{r}),F{r},IF($L$1>0,$L$5+(C{r}-$L$4)*$L$6,C{r}))",
        ).number_format = "0.00"
        plan.cell(row=r, column=8, value="=" + clk(f"G{r}"))
        # helper J: mile where an actual is present (for the anchor).
        plan.cell(row=r, column=10, value=f'=IF(ISNUMBER(F{r}),B{r},"")')

    # Anchor helpers (furthest station with an actual) — column L, hidden.
    Br = f"$B${first}:$B${last}"
    Cr = f"$C${first}:$C${last}"
    Fr = f"$F${first}:$F${last}"
    Jr = f"$J${first}:$J${last}"
    plan["K1"], plan["L1"] = "anchor mile", f"=MAX({Jr})"
    plan["K3"], plan["L3"] = "anchor pos", f"=IFERROR(MATCH($L$1,{Br},0),0)"
    plan["K4"], plan["L4"] = "anchor model", f"=IF($L$1>0,INDEX({Cr},$L$3),0)"
    plan["K5"], plan["L5"] = "anchor actual", f"=IF($L$1>0,INDEX({Fr},$L$3),0)"
    plan["K6"], plan["L6"] = "pace factor", "=IF($L$1>0,$L$5/MAX(0.1,$L$4),1)"
    for cl in ("I", "J", "K", "L"):
        plan.column_dimensions[cl].hidden = True
    for cl, w in {"A": 22, "E": 18, "F": 18, "D": 11, "H": 11}.items():
        plan.column_dimensions[cl].width = w
    plan.freeze_panes = "A5"
    plan.sheet_view.showGridLines = True

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_pacing_planner(
    ratio: pl.DataFrame,
    splits: pl.DataFrame,
    xwalk: pl.DataFrame,
    stations: pl.DataFrame,
    params: dict,
) -> str:
    """Render the standalone, mobile-first race-day pacing planner page.

    Reuses _planner_payload, derives the finish-cohort arrival-fraction table
    (_planner_eta_table), and injects a slimmed planner-only JSON plus a
    base64-embedded .xlsx (same live logic, for offline download) into
    planner_template.html.

    Inputs: es_interval_ratio, es_splits_all, es_station_xwalk,
        es_course_stations, params:reporting
    Outputs: es_pacing_planner (text HTML, data/08_reporting)
    """
    planner = _planner_payload(ratio, splits, xwalk, stations, params)
    eta = _planner_eta_table(planner)
    fmin, fmax = int(planner["fhr_min"]), int(planner["fhr_max"])
    # The client planner reads only stations + the arrival-fraction table; drop
    # the scatter points, histogram, and raw trend/avg to keep the page lean.
    slim = {
        "planner": {
            "stations": planner["stations"],
            "eta": eta,
            "fhr_min": fmin,
            "fhr_max": fmax,
        }
    }

    template = (Path(__file__).parent / "planner_template.html").read_text(
        encoding="utf-8"
    )
    marker = "/*__DATA__*/null"
    if marker not in template or "__XLSX_B64__" not in template:
        raise ValueError("planner template is missing a required marker")
    blob = json.dumps(slim, separators=(",", ":")).replace("</", "<\\/")
    html = template.replace(marker, blob)

    workbook = build_planner_workbook(planner["stations"], eta, fmin, fmax)
    xlsx_b64 = base64.b64encode(workbook).decode("ascii")
    html = html.replace("__XLSX_B64__", xlsx_b64)
    logger.info("Pacing planner built: %d bytes (xlsx %d B)", len(html), len(xlsx_b64))
    return html


# Finish-time cohorts for the blog scatter: ordered buckets take an ordinal
# one-hue ramp (slate-green, light -> dark = faster -> slower finish),
# validated with the dataviz palette checks against the theme panel.
COHORTS = [
    ("under 26 h", 0, 26, "#8aa89b"),
    ("26–30 h", 26, 30, "#5f8272"),
    ("30–34 h", 30, 34, "#3d5a50"),
    ("34 h and over", 34, 99, "#233a31"),
]


def plot_blog_interval_ratio(ratio: pl.DataFrame) -> dict:
    """Blog-ready scatter: per-leg pace vs final overall pace over the course.

    One point per finisher × aid station, all years. Each point is the runner's
    final overall pace divided by the moving pace on the leg *into* that station
    (a speed ratio): above 1.0 = that leg ran faster than their whole-race
    average, below 1.0 = a slower/harder leg. Per-cohort median lines use the
    2026 station mile marks so all years share an x position and trace the
    course's difficulty profile.

    Inputs: es_interval_ratio
    Outputs: es_blog_figures (PNG), es_blog_figures_svg (SVG)
    """
    mpl_theme.apply()
    fig, ax = plt.subplots(figsize=(12, 7))

    # Speed ratio (final pace ÷ leg pace), inverted per-row before aggregating.
    ratio = ratio.with_columns(_speed_ratio_expr("interval_ratio").alias("_speed"))

    for label, lo, hi, color in COHORTS:
        cohort = ratio.filter(
            (pl.col("finish_elapsed_hrs") >= lo) & (pl.col("finish_elapsed_hrs") < hi)
        )
        ax.scatter(
            cohort["as_dist_from_start"],
            cohort["_speed"],
            s=7,
            color=color,
            alpha=0.22,
            linewidths=0,
            label=None,
        )
        medians = (
            cohort.filter(pl.col("station_2026").is_not_null())
            .group_by("station_2026", "station_mi_2026")
            .agg(pl.col("_speed").median().alias("med"), pl.len().alias("n"))
            .filter(pl.col("n") >= MIN_GROUP_N)
            .sort("station_mi_2026")
        )
        ax.plot(
            medians["station_mi_2026"],
            medians["med"],
            color=color,
            linewidth=2,
            label=label,
            solid_capstyle="round",
        )

    ax.axhline(
        1.0, color=mpl_theme.COLORS["range"], linewidth=1.2, linestyle=(0, (4, 3))
    )
    ax.text(
        1.0,
        1.004,
        "1.0 = your final overall pace",
        fontsize=9,
        color=mpl_theme.COLORS["tick"],
        va="bottom",
    )

    lo_y = max(0.4, ratio["_speed"].quantile(0.01) - 0.03)
    hi_y = min(2.4, ratio["_speed"].quantile(0.99) + 0.03)
    ax.set_xlim(0, 106)
    ax.set_ylim(lo_y, hi_y)
    ax.legend(title=None, loc="upper left", markerscale=1.5)
    mpl_theme.set_title(
        ax,
        "Which legs make you pay",
        "Leg speed relative to final overall pace — finishers, 2016–2025",
    )
    mpl_theme.set_labels(ax, "Distance from start [mi]", "Speed relative to final")

    return (
        {"interval_ratio_scatter.png": fig},
        {"interval_ratio_scatter.svg": fig},
    )
